import csv
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from credit_card_extractor.exporters import export_csv, export_xlsx
from credit_card_extractor.i18n import Language


class TestExportCsv:
    def test_creates_file(self, sample_result, tmp_path):
        out = tmp_path / "out.csv"
        export_csv(sample_result, out)
        assert out.exists()

    def test_english_headers(self, sample_result, tmp_path):
        out = tmp_path / "out.csv"
        export_csv(sample_result, out, language=Language.EN)
        with out.open(encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            headers = next(reader)
        assert headers == ["Date", "Description", "Amount"]

    def test_portuguese_headers(self, sample_result, tmp_path):
        out = tmp_path / "out.csv"
        export_csv(sample_result, out, language=Language.PT)
        with out.open(encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            headers = next(reader)
        assert headers == ["Data", "Descrição", "Valor"]

    def test_utf8_bom(self, sample_result, tmp_path):
        out = tmp_path / "out.csv"
        export_csv(sample_result, out)
        raw = out.read_bytes()
        assert raw[:3] == b"\xef\xbb\xbf"

    def test_row_count(self, sample_result, tmp_path):
        out = tmp_path / "out.csv"
        export_csv(sample_result, out)
        with out.open(encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        # header + 3 transactions
        assert len(rows) == 4

    def test_date_format(self, sample_result, tmp_path):
        out = tmp_path / "out.csv"
        export_csv(sample_result, out)
        with out.open(encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            next(reader)  # skip header
            first_row = next(reader)
        assert first_row[0] == "2024-03-15"

    def test_negative_amount_as_string(self, sample_result, tmp_path):
        out = tmp_path / "out.csv"
        export_csv(sample_result, out)
        with out.open(encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            next(reader)
            first_row = next(reader)
        assert first_row[2] == "-125.40"

    def test_empty_result_writes_only_header(self, tmp_path, sample_result):
        from credit_card_extractor.models import ExtractionResult

        empty = ExtractionResult(
            transactions=[],
            warnings=[],
            source_file="test.pdf",
            pages_processed=0,
        )
        out = tmp_path / "empty.csv"
        export_csv(empty, out)
        with out.open(encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        assert len(rows) == 1  # header only


class TestExportXlsx:
    def test_creates_file(self, sample_result, tmp_path):
        out = tmp_path / "out.xlsx"
        export_xlsx(sample_result, out)
        assert out.exists()

    def test_file_is_valid_xlsx(self, sample_result, tmp_path):
        out = tmp_path / "out.xlsx"
        export_xlsx(sample_result, out)
        wb = openpyxl.load_workbook(str(out))
        assert "Transactions" in wb.sheetnames

    def test_english_headers(self, sample_result, tmp_path):
        out = tmp_path / "out.xlsx"
        export_xlsx(sample_result, out, language=Language.EN)
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Transactions"]
        headers = [ws.cell(row=1, column=i + 1).value for i in range(3)]
        assert headers == ["Date", "Description", "Amount"]

    def test_portuguese_headers(self, sample_result, tmp_path):
        out = tmp_path / "out.xlsx"
        export_xlsx(sample_result, out, language=Language.PT)
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Transactions"]
        headers = [ws.cell(row=1, column=i + 1).value for i in range(3)]
        assert headers == ["Data", "Descrição", "Valor"]

    def test_header_is_bold(self, sample_result, tmp_path):
        out = tmp_path / "out.xlsx"
        export_xlsx(sample_result, out)
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Transactions"]
        for cell in ws[1]:
            assert cell.font.bold is True

    def test_amount_is_numeric(self, sample_result, tmp_path):
        out = tmp_path / "out.xlsx"
        export_xlsx(sample_result, out)
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Transactions"]
        # Row 2, column 3 (amount)
        cell = ws.cell(row=2, column=3)
        assert isinstance(cell.value, float)

    def test_row_count(self, sample_result, tmp_path):
        out = tmp_path / "out.xlsx"
        export_xlsx(sample_result, out)
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Transactions"]
        assert ws.max_row == 4  # header + 3 transactions
