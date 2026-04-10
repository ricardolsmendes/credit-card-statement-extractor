import csv
from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from credit_card_extractor.i18n import DEFAULT_COLUMNS, Language, get_headers
from credit_card_extractor.models import ExtractionResult, Transaction


def _transaction_to_row(t: Transaction, columns: list[str]) -> list:
    """Map Transaction fields to an ordered list matching the given columns."""
    mapping: dict[str, object] = {
        "date": t.date.strftime("%Y-%m-%d"),
        "description": t.description,
        "amount": t.amount,
        "reference": t.reference,
        "category": t.category,
    }
    return [mapping[col] for col in columns]


def export_csv(
    result: ExtractionResult,
    output_path: Path,
    language: Language = Language.EN,
    columns: list[str] = DEFAULT_COLUMNS,
) -> None:
    """
    Write transactions to CSV.
    Uses UTF-8 with BOM (utf-8-sig) so Excel on Windows opens it correctly,
    including special characters like Descrição.
    """
    headers = get_headers(language, columns)
    with output_path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        for t in result.transactions:
            row = _transaction_to_row(t, columns)
            # Convert Decimal to string for CSV to avoid scientific notation
            writer.writerow([str(v) if isinstance(v, Decimal) else v for v in row])


def export_xlsx(
    result: ExtractionResult,
    output_path: Path,
    language: Language = Language.EN,
    columns: list[str] = DEFAULT_COLUMNS,
) -> None:
    """
    Write transactions to Excel (.xlsx) with formatted header row and
    numeric amount column.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = get_headers(language, columns)

    # Write header row with bold + light-blue fill
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # Write data rows
    for t in result.transactions:
        row = _transaction_to_row(t, columns)
        ws.append(row)

    # Apply number format to amount column (cast Decimal -> float for openpyxl)
    if "amount" in columns:
        amount_col_idx = columns.index("amount") + 1  # 1-based
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=amount_col_idx)
            cell.number_format = "#,##0.00"
            if isinstance(cell.value, Decimal):
                cell.value = float(cell.value)

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(str(output_path))
